import time, re, csv, glob, pprint, time, getpass
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException, StaleElementReferenceException, TimeoutException
from lib.stopwatch import Timer
from pyexcel.cookbook import merge_all_to_a_book
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.styles import Font, Fill, PatternFill

##################### Web scrape website #####################

timer = Timer()
timer.start()
pp = pprint.PrettyPrinter(indent=4)

# Open firefox browser
driver = webdriver.Firefox()

# Navigate to URL
driver.get("https://lds.org")

# Navigate to login
account_button = driver.find_element_by_xpath("//*[contains(text(), 'My Account')]")
account_button.click()
signin_button = driver.find_element_by_xpath("//*[contains(text(), 'Sign In')]")
signin_button.click()


# Prompt for login info
username = input('Enter username: ')
password = getpass.getpass('Enter password: ')
WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, 'IDToken1')))

# Grab inputs
username_input = driver.find_element_by_name("IDToken1")
password_input = driver.find_element_by_name("IDToken2")

# Enter values
username_input.clear()
username_input.send_keys(username)
password_input.clear()
password_input.send_keys(password)

# Submit form
username_input.submit()

# Wait for browser to refresh
WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, 'pf-signin')))

# Navigate to directory
account_button = driver.find_element_by_xpath("//*[contains(text(), 'My Account')]")
account_button.click()
directory_button = driver.find_element_by_xpath("//*[contains(text(), 'Directory')]")
directory_button.click()

## Grab information and put into dictionary
households = []
counter1 = 0
household_link = WebDriverWait(driver, 10).until(EC.presence_of_element_located((
    By.ID, 'listItem0')))

while(household_link):
    # Grab next household link based on counter
    household_id = 'listItem' + str(counter1)
    try:
        # Grab link and household name
        household_link = driver.find_element_by_id(household_id) \
            .find_element_by_tag_name('a')
        household_name = household_link.text

        # Click into household
        household_link.click()

        # Wait for ajox to load then assign to element
        element = WebDriverWait(driver, 40).until(
            EC.presence_of_element_located((By.ID, 'householdAddress'))
            )

        # Grab household address
        household_address_spans = element.find_elements_by_tag_name('span')

        # Change final_pos if household is the logged in user
        try:
            driver.find_element_by_id("show_profile_edit")
            final_pos = (len(household_address_spans) - 2)

         # Assign for all other households
        except NoSuchElementException:
            final_pos = (len(household_address_spans) - 1)

        i = 0
        household_street_address = ''
        household_city = ''
        household_zip = ''

        # Loop to grab the street address
        while(i < final_pos):
            household_street_address += ' {}'.format(household_address_spans[i].text)
            i += 1

        # Grab city
        household_city = household_address_spans[final_pos].text.split(',')[0]

        # Grab zip
        household_zip = re.search(
            r'(\d{5})',
            household_address_spans[final_pos].text
            )
        if household_zip:
            household_zip = household_zip.group(0)
        else:
            household_zip = '84606'

        # Add to list
        households.append((
            '',
            household_name,
            household_street_address.replace(".", "").strip(),
            household_city,
            'United States',
            household_zip,
            'Active Member',
            '',
            ))

        # Show progress of homes
        for value in households:
            print('\n{}'.format(value[1]))
            print(value[2])
            print(value[3])
            print(value[5])

    # Output when finished
    except NoSuchElementException:
        print("\nFinished compiling each household.")
        household_link = None
        print(timer.stop(message = 'Time Elapsed: '))

    # Increment
    counter1 += 1

## Write to csv file
with open('data/backup/map_data_backup.csv', 'w', newline='') as csvfile:
    fieldnames = [
        'Location Code',
        'Name',
        'Street',
        'City',
        'Country',
        'Zipcode',
        'Category',
        'Description',]
    writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
    writer.writeheader()

    # Write each household to individual rows
    for value in households:
        writer.writerow({
            'Location Code': value[0],
            'Name': value[1],
            'Street': value[2],
            'City': value[3],
            'Country': value[4],
            'Zipcode': value[5],
            'Category': value[6],
            'Description': value[7],
            })

# Convert csv file to xlsx
merge_all_to_a_book(glob.glob("data/backup/map_data_backup.csv"), "data/map_data.xlsx")

### For testing purposes ###
households = []
# Load up new map data
wb = load_workbook(filename='data/map_data.xlsx', read_only=True)
ws = wb['map_data_backup.csv']

for index, row in enumerate(ws.rows):
    if (index != 0):
        households.append((
            row[0].value,
            row[1].value,
            row[2].value,
            row[3].value,
            row[4].value,
            row[5].value,
            row[6].value,
            row[7].value,
            ))

wb.close()
### End test block ###

##################### Compare map excel file data #####################
current_households = []
misc_households = []

upload_households = []
change_households = []
add_households = []
delete_households = []

## Load up current household information
##TODO will have to change worksheet on official map
wb = load_workbook(filename='data/current_map_data.xlsm', read_only=True)
ws = wb.worksheets[0]

for index, row in enumerate(ws.rows):
    # Skip header info
    if (index != 0):
        # Add to the normal list if a normal member
        if row[6].value == 'Less Active Member' or \
            row[6].value == 'Active Member' or \
            row[6].value == 'Part Member Family':

            current_households.append((
                row[0].value,
                row[1].value,
                row[2].value,
                row[3].value,
                row[4].value,
                row[5].value,
                row[6].value,
                row[7].value,
                ))

        # Put other households in a misc list that will be added to the upload list at the end
        else:
            misc_households.append([
                row[0].value,
                row[1].value,
                row[2].value,
                row[3].value,
                row[4].value,
                row[5].value,
                row[6].value,
                row[7].value,
                ])

wb.close()

### Add household to upload lists if name and match in current households
for value in households:
    is_current_household = False
    is_misc_household = False

    # Check if household name and address is in the current household
    for current_value in current_households:
        # If the name and address already matches, add current household to upload list
        if value[1] == current_value[1] and value[2] == current_value[2]:
            upload_households.append(current_value)
            is_current_household = True

    # Check if address is in a prior household
    if is_current_household == False:
        for misc_value in misc_households:

            # If address is the same as a prior household, update the prior household and add to upload list
            if value[2] == misc_value[2]:
                misc_value[1] = value[1]
                misc_value[6] = value[6]
                misc_value[7] = value[7]

                change_households.append(misc_value)
                is_misc_household = True

    # Upload any of the remaining new households
    if is_current_household == False and is_misc_household == False:
        # Change color to indicate a new household
        add_households.append(value)

## Find households to be deleted
upload_misc_households = []

for misc_value in misc_households:
    no_match = True

    # Find if prior residencies need to be deleted
    for upload_value in upload_households:
        if misc_value[2] == upload_value[2]:
            delete_households.append(misc_value)
            no_match = False

    # Add to official misc list if there was no match
    if no_match:
        upload_misc_households.append(misc_value)

## Upload to official upload excel sheet
pp.pprint(add_households)
print('*' * 80)
pp.pprint(change_households)
print('*' * 80)
pp.pprint(delete_households)
print('*' * 80)

csv_counter = 0
add_counter = 0
change_counter = 0
delete_counter = 0

with open('data/backup/updated_map_data_backup.csv', 'w', newline='') as csvfile:
    fieldnames = [
        'Location Code',
        'Name',
        'Street',
        'City',
        'Country',
        'Zipcode',
        'Category',
        'Description',]
    writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
    writer.writeheader()

    # Write each household to individual rows
    for value in add_households:
        writer.writerow({
            'Location Code': value[0],
            'Name': value[1],
            'Street': value[2],
            'City': value[3],
            'Country': value[4],
            'Zipcode': value[5],
            'Category': value[6],
            'Description': value[7],
            })
        add_counter += 1
        csv_counter += 1


    for value in change_households:
        writer.writerow({
            'Location Code': value[0],
            'Name': value[1],
            'Street': value[2],
            'City': value[3],
            'Country': value[4],
            'Zipcode': value[5],
            'Category': value[6],
            'Description': value[7],
            })
        change_counter += 1
        csv_counter += 1

    for value in delete_households:
        writer.writerow({
            'Location Code': value[0],
            'Name': value[1],
            'Street': value[2],
            'City': value[3],
            'Country': value[4],
            'Zipcode': value[5],
            'Category': value[6],
            'Description': value[7],
            })
        delete_counter += 1
        csv_counter += 1

    for value in upload_households:
        writer.writerow({
            'Location Code': value[0],
            'Name': value[1],
            'Street': value[2],
            'City': value[3],
            'Country': value[4],
            'Zipcode': value[5],
            'Category': value[6],
            'Description': value[7],
            })
        csv_counter += 1

    for value in upload_misc_households:
        writer.writerow({
            'Location Code': value[0],
            'Name': value[1],
            'Street': value[2],
            'City': value[3],
            'Country': value[4],
            'Zipcode': value[5],
            'Category': value[6],
            'Description': value[7],
            })
        csv_counter += 1

# Format to xlsx file
merge_all_to_a_book(glob.glob("data/backup/updated_map_data_backup.csv"), "data/updated_map_data.xlsx")

## Format colors for add, change, and delete households

# Load updated excel book
wb = load_workbook(filename='data/updated_map_data.xlsx')
ws = wb['updated_map_data_backup.csv']

if add_counter != 0:
    # Load color fills for each added households
    fill = PatternFill(patternType='solid', fgColor='6bf977')
    start_cell = 'A' + str(2)
    end_cell = 'H' + str(1 + add_counter)

    for row in ws[start_cell : end_cell]:
        for cell in row:
            cell.fill = fill

if change_counter != 0:
    # Load color fills for each changed households
    fill = PatternFill(patternType='solid', fgColor='faff77')
    start_cell = 'A' + str(2 + add_counter)
    end_cell = 'H' + str(1 + add_counter + change_counter)

    for row in ws[start_cell : end_cell]:
        for cell in row:
            cell.fill = fill

if delete_counter != 0:
    # Load color fills for each changed households
    fill = PatternFill(patternType='solid', fgColor='f96f57')
    start_cell = 'A' + str(2 + add_counter + change_counter)
    end_cell = 'H' + str(1 + add_counter + change_counter + delete_counter)

    for row in ws[start_cell : end_cell]:
        for cell in row:
            cell.fill = fill


print(add_counter)
print(change_counter)
print(delete_counter)
print(csv_counter)

# Save with color updates
wb.save('data/updated_map_data.xlsx')

print('\nThe pending changes can now be seen in the "updated_map_data.xlx" file.')
print('To continue, press any key...')
input()
##################### Upload to Zeemaps #####################

### For testing purposes ###
# driver = webdriver.Firefox()
### End test block ###

# # Navigate to URL
# driver.get("https://www.zeemaps.com/map?group=2991393")
#
# # Navigate to login
# signin_button = driver.find_element_by_xpath("//*[contains(text(), 'Sign-in')]")
# signin_button.click()
#
# # Prompt for login info
# # username = input('Enter username:')
# # password = input('Enter password:')
# WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, 'u_email')))
# username = 'pdrean4@gmail.com'
# password = 'Znephite44'
#
#
# # Grab inputs
# username_input = driver.find_element_by_name("u_email")
# password_input = driver.find_element_by_name("u_passwd")
# submit_button = driver.find_element_by_xpath("//*[contains(text(), 'Sign-In')]")
#
# # Submit login
# username_input.send_keys(username)
# password_input.send_keys(password)
#
# submit_button.click()
#
# WebDriverWait(driver, 30).until(EC.presence_of_element_located(
#     (By.XPATH, '//button[@title = "Zoom out"]')
#     ))
#
# # Open navigation bar see map markers
# navigation_button = driver.find_element_by_class_name('ui-layout-toggler')
# navigation_button.click()
#
# # Load each household element into a list
# web_delete_households = driver.find_element_by_id('listofentries') \
#     .find_elements_by_tag_name('td')
#
# # Get length
# data_length = (len(driver.find_element_by_id('listofentries') \
#     .find_elements_by_tag_name('td')) + 1)
#
# print(data_length)
# # Loop through each household to delete it
# for thing in range(0, data_length):
#     try:
#         item = WebDriverWait(driver, 5).until(EC.presence_of_element_located(
#             (By.TAG_NAME, 'td')
#             ))
#         print(item.text)
#         item.click()
#
#         trash_button = WebDriverWait(driver, 30).until(EC.presence_of_element_located(
#             (By.CLASS_NAME, 'iwtrash')
#             ))
#         trash_button.click()
#
#         yes_button = WebDriverWait(driver, 30).until(EC.presence_of_element_located(
#             (By.XPATH, "//*[contains(text(), 'Yes')]")
#             ))
#         yes_button.click()
#     except StaleElementReferenceException:
#         pass
#     except TimeoutException:
#         pass
#
# print('\nAll households have been deleted.')
# print('Adding new households now...')
#
# ### Add the new households one by one
# for add_value in add_households:
#
#     # Navigate to adding markers
#     addition_button = WebDriverWait(driver, 30).until(EC.presence_of_element_located(
#         (By.XPATH, "//*[contains(text(), 'Additions')]")
#         ))
#     addition_button.click()
#
#     add_marker_button = WebDriverWait(driver, 30).until(EC.presence_of_element_located(
#         (By.XPATH, "//*[contains(text(), 'Add Marker - Simple')]")
#         ))
#     add_marker_button.click()
#
#     # Grab inputs
#     name_input = WebDriverWait(driver, 30).until(EC.presence_of_element_located(
#         (By.NAME, "attr(name)")))
#     location_input = driver.find_element_by_name("location")
#
#     # Fill with values
#     name_input.send_keys(add_value[1])
#     location_input.send_keys('{} {} {}'.format(add_value[2], add_value[3], add_value[5]))
#
#     # Make choice on select lists
#     select = Select(driver.find_element_by_name('attr(color)'))
#     select.select_by_index(1)
#
#     # Submit addition
#     submit_button = WebDriverWait(driver, 30).until(EC.presence_of_element_located(
#         (By.XPATH, "//*[contains(text(), 'Submit')]")))
#     print(driver.find_element_by_class_name('ui-button'))
#     print(driver.find_element_by_class_name('ui-button').text)
#     driver.find_element_by_class_name('ui-button').click()
#     print(submit_button)
#     print(submit_button.text)
#     # submit_button.click()
#
#
#     # # Confirm addition
#     # ok_button = WebDriverWait(driver, 30).until(EC.presence_of_element_located(
#     #     (By.XPATH, "//*[contains(text(), 'OK')]")))
#     # ok_button = driver.find_element_by_xpath("//*[contains(text(), 'OK')]").find_element_by_xpath('..')
#     # ok_button.submit()


print('\nPress any button to  quit.')
# Quit process
input()
driver.quit()
