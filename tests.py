import unittest
# from map_z_scraper import *
from openpyxl import load_workbook

class MapTestCast(unittest.TestCase):
    def __init__(self, *args, **kwargs):
        self.new_households = {}
        self.current_households = {}
        self.misc_households = {}
        super().__init__(*args, **kwargs)

    def setUp(self):
        '''Load up new map data and current map data into two lists'''
        # Load up new map data
        wb = load_workbook(filename='data/map_data.xlsx', read_only=True)
        ws = wb['map_data_backup.csv']

        for index, row in enumerate(ws.rows):
            if (index != 0):
                self.new_households[row[0].value] = (
                    row[1].value,
                    row[2].value,
                    row[3].value,
                    )
        wb.close()

        # Load up current map data
        #TODO change name of xl sheet
        wb = load_workbook(filename='data/current_map_data.xlsx', read_only=True)
        ws = wb['compiled_data_1.2']

        for index, row in enumerate(ws.rows):
            # Skip header info
            if (index != 0):
                if row[6].value == 'Less Active Member' or \
                    row[6].value == 'Active Member' or \
                    row[6].value == 'Part Member Family':

                    self.current_households[row[1].value] = (
                        row[2].value,
                        row[3].value,
                        row[4].value,
                        row[5].value,
                        row[6].value,
                        row[7].value,
                        row[0].value,
                        )
                    print(self.current_households[row[1].value])
                else:
                    self.misc_households[row[1].value] = (
                        row[2].value,
                        row[3].value,
                        row[4].value,
                        row[5].value,
                        row[6].value,
                        row[7].value,
                        row[0].value,
                        )
                    print('MISC: ' + str(self.misc_households[row[1].value]))
        wb.close()

    def test_decimal(self):
        '''Test there's no decimals in the household street address'''
        no_decimals = True
        for key, value in self.new_households.items():
            if '.' in value[0]:
                no_decimals = False
        self.assertTrue(no_decimals)

    def test_names(self):
        '''Make sure household names and formats generally matchup'''
        x = self.new_households['Adams, Lola']
        y = self.current_households['Adams, Lola']

        self.assertEqual(x[0],y[0])

if __name__ == '__main__':
    unittest.main()
